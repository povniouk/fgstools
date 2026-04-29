import os
import re
import json
import datetime
import sqlite3
import tempfile
from collections import defaultdict
import openpyxl
from flask import Blueprint, request, jsonify

bp = Blueprint('spi', __name__)

DB_PATH = os.environ.get("DB_PATH", "cwlng.db")
_log = print

FGS_SYSTEMS = {'FGS', 'LFGS'}

_COLS = [
    'tag_number', 'tag_type', 'system1', 'io_type1', 'typical',
    'tag_serv', 'area_class', 'unit_name', 'design_by', 'status',
    'loop_name', 'plant_area', 'instr_type', 'instr_desc',
    'area', 'cwa', 'ex_type', 'signal_type', 'io_type2', 'system2',
    'fgs_via_system2',  # flag: True when FGS assignment is in System2
]

_COL_MAP = {
    'tag_number': 'Tag_Number', 'tag_type': 'Tag_Type', 'system1': 'System1',
    'io_type1': 'IO_Type1', 'typical': 'Typical', 'tag_serv': 'Tag_Serv',
    'area_class': 'Area_Class', 'unit_name': 'Unit_name', 'design_by': 'Design_By',
    'status': 'Status', 'loop_name': 'Loop_Name', 'plant_area': 'Plant_Area',
    'instr_type': 'Instr_Type', 'instr_desc': 'Instr_Desc',
    'area': 'Area', 'cwa': 'CWA', 'ex_type': 'Ex_Type',
    'signal_type': 'Signal_Type', 'io_type2': 'IO_Type2', 'system2': 'System2',
}


def get_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys = ON")
    return db


def init_db():
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS spi_imports (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            filename    TEXT,
            week_label  TEXT,
            imported_at TEXT,
            total_rows  INTEGER,
            fgs_count   INTEGER
        );
        CREATE TABLE IF NOT EXISTS spi_tags (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            import_id       INTEGER REFERENCES spi_imports(id) ON DELETE CASCADE,
            tag_number      TEXT,
            tag_type        TEXT,
            system1         TEXT,
            io_type1        TEXT,
            typical         TEXT,
            tag_serv        TEXT,
            area_class      TEXT,
            unit_name       TEXT,
            design_by       TEXT,
            status          TEXT,
            loop_name       TEXT,
            plant_area      TEXT,
            instr_type      TEXT,
            instr_desc      TEXT,
            area            TEXT,
            cwa             TEXT,
            ex_type         TEXT,
            signal_type     TEXT,
            io_type2        TEXT,
            system2         TEXT,
            fgs_via_system2 INTEGER DEFAULT 0
        );
        CREATE INDEX IF NOT EXISTS idx_spi_tags_import  ON spi_tags(import_id);
        CREATE INDEX IF NOT EXISTS idx_spi_tags_system1 ON spi_tags(system1);
        CREATE INDEX IF NOT EXISTS idx_spi_tags_system2 ON spi_tags(system2);
        CREATE TABLE IF NOT EXISTS spi_diffs (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            import_id      INTEGER REFERENCES spi_imports(id) ON DELETE CASCADE,
            prev_import_id INTEGER,
            computed_at    TEXT,
            new_count      INTEGER DEFAULT 0,
            removed_count  INTEGER DEFAULT 0,
            changed_count  INTEGER DEFAULT 0,
            diff_json      TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_spi_diffs_import ON spi_diffs(import_id);
    """)
    db.commit()
    db.close()


def _extract_week(filename):
    m = re.search(r'W(\d+)', filename, re.IGNORECASE)
    return f"W{m.group(1)}" if m else "W?"


def _val(row, idx):
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None:
        return None
    s = str(v).strip()
    return None if s in ('-', '', 'None') else s


def parse_spi_excel(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    headers = None
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row[:10]):
            headers = row
            break
    if not headers:
        return [], 0

    h = {str(v): i for i, v in enumerate(headers) if v is not None}

    tags = []
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None and str(v).strip() not in ('', '-') for v in row[:20]):
            continue
        total += 1

        sys1 = _val(row, h.get('System1'))
        sys2 = _val(row, h.get('System2'))
        fgs_via_sys2 = sys1 not in FGS_SYSTEMS and sys2 in FGS_SYSTEMS

        if sys1 not in FGS_SYSTEMS and not fgs_via_sys2:
            continue

        tag = {col: _val(row, h.get(_COL_MAP[col])) for col in _COL_MAP}
        tag['fgs_via_system2'] = 1 if fgs_via_sys2 else 0

        if tag['tag_number']:
            tag['tag_number'] = re.sub(r'\s+', ' ', tag['tag_number']).strip()

        tags.append(tag)

    return tags, total


def _compute_flags(tag):
    flags = []
    if not tag.get('typical'):
        flags.append('missing_typical')
    if not tag.get('area_class'):
        flags.append('missing_area_class')
    if not tag.get('tag_type'):
        flags.append('missing_tag_type')
    if tag.get('status') == 'TBF':
        flags.append('status_tbf')
    if tag.get('fgs_via_system2'):
        flags.append('via_system2')
    return flags


def _duplicate_tag_numbers(db, import_id):
    """Returns the set of tag_numbers that appear more than once in this import."""
    rows = db.execute(
        "SELECT tag_number FROM spi_tags WHERE import_id=? AND tag_number IS NOT NULL "
        "GROUP BY tag_number HAVING COUNT(*) > 1",
        (import_id,)
    ).fetchall()
    return {r['tag_number'] for r in rows}


_DIFF_FIELDS = [
    'typical', 'status', 'area_class',
    'system1', 'io_type1', 'system2', 'io_type2',
    'design_by', 'loop_name', 'tag_serv',
]

# Display-relevant fields stored on new/removed tag snapshots
_DIFF_SNAPSHOT_FIELDS = [
    'tag_number', 'loop_name', 'system1', 'io_type1',
    'system2', 'io_type2', 'typical', 'tag_serv', 'area_class', 'design_by', 'status',
]


def _compute_diff(new_tags, prev_tags):
    """Compare two tag lists keyed on tag_number. Returns a diff dict."""
    new_map  = {t['tag_number']: t for t in new_tags  if t.get('tag_number')}
    prev_map = {t['tag_number']: t for t in prev_tags if t.get('tag_number')}

    new_keys  = set(new_map)
    prev_keys = set(prev_map)

    added = [
        {f: new_map[k].get(f) for f in _DIFF_SNAPSHOT_FIELDS}
        for k in sorted(new_keys - prev_keys)
    ]
    removed = [
        {f: prev_map[k].get(f) for f in _DIFF_SNAPSHOT_FIELDS}
        for k in sorted(prev_keys - new_keys)
    ]

    changed = []
    for k in sorted(new_keys & prev_keys):
        n, p = new_map[k], prev_map[k]
        field_changes = {}
        for f in _DIFF_FIELDS:
            nv = n.get(f) or ''
            pv = p.get(f) or ''
            if nv != pv:
                field_changes[f] = {'from': p.get(f), 'to': n.get(f)}
        if field_changes:
            changed.append({
                'tag_number': k,
                'loop_name': n.get('loop_name'),
                'changes': field_changes,
            })

    return {
        'new': added, 'removed': removed, 'changed': changed,
        'new_count': len(added), 'removed_count': len(removed), 'changed_count': len(changed),
    }


# ── Routes ────────────────────────────────────────────────────────────────────

@bp.route('/api/spi/import', methods=['POST'])
def import_spi():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Only .xlsx files accepted'}), 400

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    week_label = _extract_week(f.filename)
    _log(f"[SPI] Importing {f.filename} ({week_label})...")

    try:
        tags, total_rows = parse_spi_excel(tmp_path)
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    now = datetime.datetime.utcnow().isoformat()
    db = get_db()
    cur = db.execute(
        "INSERT INTO spi_imports (filename, week_label, imported_at, total_rows, fgs_count) VALUES (?,?,?,?,?)",
        (f.filename, week_label, now, total_rows, len(tags))
    )
    import_id = cur.lastrowid

    insert_cols = [c for c in _COLS if c != 'fgs_via_system2'] + ['fgs_via_system2']
    db.executemany(
        f"INSERT INTO spi_tags (import_id, {', '.join(insert_cols)}) "
        f"VALUES (:import_id, {', '.join(':'+c for c in insert_cols)})",
        [{**t, 'import_id': import_id} for t in tags]
    )
    db.commit()

    flag_counts = {}
    for t in tags:
        for flag in _compute_flags(t):
            flag_counts[flag] = flag_counts.get(flag, 0) + 1

    # Diff against the most recent previous import
    diff_summary = None
    prev = db.execute(
        "SELECT id, week_label FROM spi_imports WHERE id != ? ORDER BY id DESC LIMIT 1",
        (import_id,)
    ).fetchone()
    if prev:
        prev_tags = [dict(r) for r in db.execute(
            "SELECT * FROM spi_tags WHERE import_id=?", (prev['id'],)
        ).fetchall()]
        diff = _compute_diff(tags, prev_tags)
        db.execute(
            "INSERT INTO spi_diffs "
            "(import_id, prev_import_id, computed_at, new_count, removed_count, changed_count, diff_json) "
            "VALUES (?,?,?,?,?,?,?)",
            (import_id, prev['id'], now,
             diff['new_count'], diff['removed_count'], diff['changed_count'],
             json.dumps(diff))
        )
        db.commit()
        diff_summary = {
            'prev_import_id': prev['id'],
            'prev_week_label': prev['week_label'],
            'new_count': diff['new_count'],
            'removed_count': diff['removed_count'],
            'changed_count': diff['changed_count'],
        }
        _log(f"[SPI] Diff vs {prev['week_label']}: "
             f"+{diff['new_count']} new, -{diff['removed_count']} removed, "
             f"~{diff['changed_count']} changed")

    db.close()
    _log(f"[SPI] Done: {len(tags)} F&G tags ({flag_counts.get('via_system2',0)} via System2) "
         f"from {total_rows} total rows — flags: {flag_counts}")
    return jsonify({
        'ok': True, 'import_id': import_id, 'week_label': week_label,
        'total_rows': total_rows, 'fgs_count': len(tags), 'flag_counts': flag_counts,
        'diff': diff_summary,
    })


@bp.route('/api/spi/imports', methods=['GET'])
def list_imports():
    db = get_db()
    rows = db.execute("SELECT * FROM spi_imports ORDER BY id DESC").fetchall()
    diffs = {r['import_id']: dict(r) for r in db.execute(
        "SELECT import_id, prev_import_id, new_count, removed_count, changed_count "
        "FROM spi_diffs"
    ).fetchall()}
    # Fetch week labels for prev imports
    all_imports = {r['id']: r['week_label'] for r in
                   db.execute("SELECT id, week_label FROM spi_imports").fetchall()}
    db.close()
    result = []
    for r in rows:
        imp = dict(r)
        d = diffs.get(r['id'])
        if d:
            d['prev_week_label'] = all_imports.get(d['prev_import_id'], '?')
        imp['diff'] = d
        result.append(imp)
    return jsonify({'imports': result})


@bp.route('/api/spi/diff/<int:import_id>', methods=['GET'])
def get_diff(import_id):
    db = get_db()
    row = db.execute(
        "SELECT * FROM spi_diffs WHERE import_id=?", (import_id,)
    ).fetchone()
    db.close()
    if not row:
        return jsonify({'diff': None, 'import_id': import_id})
    d = dict(row)
    d['diff_data'] = json.loads(d['diff_json'])
    del d['diff_json']
    return jsonify({'diff': d, 'import_id': import_id})


@bp.route('/api/spi/imports/<int:import_id>', methods=['DELETE'])
def delete_import(import_id):
    db = get_db()
    row = db.execute("SELECT id FROM spi_imports WHERE id=?", (import_id,)).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Import not found'}), 404
    db.execute("DELETE FROM spi_imports WHERE id=?", (import_id,))
    db.commit()
    db.close()
    _log(f"[SPI] Deleted import {import_id}")
    return jsonify({'ok': True})


@bp.route('/api/spi/tags', methods=['GET'])
def get_tags():
    import_id  = request.args.get('import_id')
    system1    = request.args.get('system1', '')
    io_type1   = request.args.get('io_type1', '')
    design_by  = request.args.get('design_by', '')
    flags_only = request.args.get('flags_only', '')

    db = get_db()
    if not import_id:
        latest = db.execute("SELECT id FROM spi_imports ORDER BY id DESC LIMIT 1").fetchone()
        if not latest:
            db.close()
            return jsonify({'tags': [], 'import_id': None, 'count': 0})
        import_id = latest['id']

    q = "SELECT * FROM spi_tags WHERE import_id = ?"
    params = [import_id]
    if system1:
        # Match System1 OR System2 so FGS tags assigned via secondary system appear
        q += " AND (system1 = ? OR system2 = ?)"
        params.extend([system1, system1])
    if io_type1:
        q += " AND io_type1 = ?"; params.append(io_type1)
    if design_by:
        q += " AND design_by = ?"; params.append(design_by)

    tags = [dict(r) for r in db.execute(q, params).fetchall()]
    dupes = _duplicate_tag_numbers(db, import_id)
    db.close()

    for t in tags:
        t['flags'] = _compute_flags(t)
        if t.get('tag_number') in dupes:
            t['flags'].append('duplicate_tag')

    if flags_only:
        tags = [t for t in tags if t['flags']]

    return jsonify({'tags': tags, 'import_id': int(import_id), 'count': len(tags)})


@bp.route('/api/spi/loops', methods=['GET'])
def get_loops():
    import_id     = request.args.get('import_id')
    system1       = request.args.get('system1', '')
    io_type1      = request.args.get('io_type1', '')
    design_by     = request.args.get('design_by', '')
    warnings_only = request.args.get('warnings_only', '')

    db = get_db()
    if not import_id:
        latest = db.execute("SELECT id FROM spi_imports ORDER BY id DESC LIMIT 1").fetchone()
        if not latest:
            db.close()
            return jsonify({'loops': [], 'import_id': None, 'loop_count': 0, 'tag_count': 0})
        import_id = latest['id']

    q = "SELECT * FROM spi_tags WHERE import_id = ?"
    params = [import_id]
    if system1:
        q += " AND (system1 = ? OR system2 = ?)"
        params.extend([system1, system1])
    if io_type1:
        q += " AND io_type1 = ?"; params.append(io_type1)
    if design_by:
        q += " AND design_by = ?"; params.append(design_by)
    q += " ORDER BY tag_number"

    tags = [dict(r) for r in db.execute(q, params).fetchall()]
    dupes = _duplicate_tag_numbers(db, import_id)
    db.close()

    for t in tags:
        t['flags'] = _compute_flags(t)
        if t.get('tag_number') in dupes:
            t['flags'].append('duplicate_tag')

    # Group by loop_name; null/blank → '' key, sorted to end
    loops_map = defaultdict(list)
    for t in tags:
        loops_map[t.get('loop_name') or ''].append(t)

    loops = []
    for loop_name in sorted(loops_map.keys(), key=lambda k: ('\xff' if not k else k)):
        loop_tags = loops_map[loop_name]

        typicals = [t['typical'] for t in loop_tags if t.get('typical')]
        unique_typ = list(set(typicals))
        if not unique_typ:
            typical = None
        elif len(unique_typ) == 1:
            typical = unique_typ[0]
        else:
            typical = 'Mixed'

        warnings = []
        if any(not t.get('typical') for t in loop_tags):
            warnings.append('missing_typical')
        if len(unique_typ) > 1:
            warnings.append('inconsistent_typical')
        if any(not t.get('area_class') for t in loop_tags):
            warnings.append('missing_area_class')

        loops.append({
            'loop_name': loop_name or None,
            'typical': typical,
            'tag_count': len(loop_tags),
            'warnings': warnings,
            'tags': loop_tags,
        })

    # Summary counts across all (pre-warnings-filter) loops
    flag_summary = {}
    for loop in loops:
        for tag in loop['tags']:
            for f in tag.get('flags', []):
                flag_summary[f] = flag_summary.get(f, 0) + 1

    if warnings_only:
        loops = [l for l in loops if l['warnings']]

    return jsonify({
        'loops': loops,
        'import_id': int(import_id),
        'loop_count': len(loops),
        'tag_count': sum(l['tag_count'] for l in loops),
        'flag_summary': flag_summary,
    })
